#!/usr/bin/env python3
"""Build a visual Excel overview card without starting Microsoft Excel.

This is an isolated prototype, deliberately outside the Launcher UI.  It proves
the first product brick only: one workbook revision -> useful preview image ->
cache.  The same SheetModel will later feed both the thumbnail renderer and the
interactive HTML view; no browser HTML is generated here.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


RENDERER_VERSION = "fast-preview-prototype-v3-baseline-fit"
MAX_MODEL_ROWS = 2_000
MAX_MODEL_COLUMNS = 100
PREVIEW_SIZE = (1_400, 900)
THUMBNAIL_SIZE = (360, 232)
DETAIL_ROWS = 52
DETAIL_COLUMNS = 18


@dataclass(frozen=True)
class CellModel:
    row: int
    column: int
    value: str
    bold: bool
    font_size: float
    fill: str | None
    font_color: str | None
    horizontal: str | None


@dataclass(frozen=True)
class SheetModel:
    index: int
    name: str
    first_row: int
    last_row: int
    first_column: int
    last_column: int
    cells: dict[tuple[int, int], CellModel]
    column_widths: dict[int, float]
    row_heights: dict[int, float]
    merged_starts: dict[tuple[int, int], tuple[int, int]]
    merged_children: set[tuple[int, int]]

    @property
    def range_label(self) -> str:
        return f"{get_column_letter(self.first_column)}{self.first_row}:{get_column_letter(self.last_column)}{self.last_row}"

    @property
    def nonempty_cells(self) -> int:
        return len(self.cells)


def rgb(color: Any) -> str | None:
    value = str(getattr(color, "rgb", "") or "")
    if len(value) not in {6, 8}:
        return None
    if len(value) == 8 and value[:2] == "00":
        return None
    return f"#{value[-6:]}"


def source_key(path: Path, preview_size: tuple[int, int]) -> str:
    info = path.stat()
    raw = f"{path.resolve()}|{info.st_size}|{info.st_mtime_ns}|{RENDERER_VERSION}|{preview_size[0]}x{preview_size[1]}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()[:20]


def usable_value(value: Any, cached_value: Any) -> str:
    value = cached_value if cached_value is not None else value
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:220]


def find_used_bounds(sheet: Any, values_sheet: Any) -> tuple[int, int, int, int] | None:
    first_row = first_column = None
    last_row = last_column = 0
    rows = min(int(sheet.max_row or 1), MAX_MODEL_ROWS)
    columns = min(int(sheet.max_column or 1), MAX_MODEL_COLUMNS)
    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            if usable_value(sheet.cell(row, column).value, values_sheet.cell(row, column).value):
                first_row = row if first_row is None else min(first_row, row)
                first_column = column if first_column is None else min(first_column, column)
                last_row = max(last_row, row)
                last_column = max(last_column, column)
    if first_row is None or first_column is None:
        return None
    return first_row, last_row, first_column, last_column


def sheet_score(bounds: tuple[int, int, int, int] | None, sheet: Any) -> int:
    if bounds is None:
        return -1
    first_row, last_row, first_column, last_column = bounds
    density_area = max(1, (last_row - first_row + 1) * (last_column - first_column + 1))
    # Prefer a genuinely filled sheet but do not let a 2,000-row register
    # automatically outrank a concise active cover/summary sheet.
    return min(50_000, density_area) + len(sheet.merged_cells.ranges) * 25


def build_sheet_model(styles_sheet: Any, values_sheet: Any, index: int, bounds: tuple[int, int, int, int]) -> SheetModel:
    first_row, last_row, first_column, last_column = bounds
    cells: dict[tuple[int, int], CellModel] = {}
    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            style_cell = styles_sheet.cell(row, column)
            text = usable_value(style_cell.value, values_sheet.cell(row, column).value)
            if not text:
                continue
            cells[(row, column)] = CellModel(
                row=row,
                column=column,
                value=text,
                bold=bool(style_cell.font.bold),
                font_size=float(style_cell.font.sz or 11.0),
                fill=rgb(style_cell.fill.fgColor) if style_cell.fill.fill_type == "solid" else None,
                font_color=rgb(style_cell.font.color),
                horizontal=style_cell.alignment.horizontal,
            )
    column_widths = {
        column: float(styles_sheet.column_dimensions[get_column_letter(column)].width or 8.43)
        for column in range(first_column, last_column + 1)
    }
    row_heights = {
        row: float(styles_sheet.row_dimensions[row].height or 15.0)
        for row in range(first_row, last_row + 1)
    }
    merged_starts: dict[tuple[int, int], tuple[int, int]] = {}
    merged_children: set[tuple[int, int]] = set()
    for merged in styles_sheet.merged_cells.ranges:
        if merged.max_row < first_row or merged.min_row > last_row or merged.max_col < first_column or merged.min_col > last_column:
            continue
        start = (max(merged.min_row, first_row), max(merged.min_col, first_column))
        merged_starts[start] = (min(merged.max_row, last_row) - start[0] + 1, min(merged.max_col, last_column) - start[1] + 1)
        for row in range(start[0], min(merged.max_row, last_row) + 1):
            for column in range(start[1], min(merged.max_col, last_column) + 1):
                if (row, column) != start:
                    merged_children.add((row, column))
    return SheetModel(index, styles_sheet.title, first_row, last_row, first_column, last_column, cells, column_widths, row_heights, merged_starts, merged_children)


def select_primary_sheet(styles_book: Any, values_book: Any) -> SheetModel:
    candidates: list[tuple[int, SheetModel]] = []
    active_index = styles_book.index(styles_book.active)
    for index, styles_sheet in enumerate(styles_book.worksheets):
        if styles_sheet.sheet_state != "visible":
            continue
        bounds = find_used_bounds(styles_sheet, values_book.worksheets[index])
        if bounds is None:
            continue
        model = build_sheet_model(styles_sheet, values_book.worksheets[index], index, bounds)
        # An active non-empty sheet is usually deliberately chosen by the author.
        priority = 1_000_000 if index == active_index else 0
        candidates.append((priority + sheet_score(bounds, styles_sheet), model))
    if not candidates:
        raise ValueError("В книге нет видимого листа с данными")
    return max(candidates, key=lambda item: item[0])[1]


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    filename = "segoeuib.ttf" if bold else "segoeui.ttf"
    path = Path("C:/Windows/Fonts") / filename
    if path.exists():
        return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def trim_text(draw: ImageDraw.ImageDraw, text: str, available: int, current_font: Any) -> str:
    if draw.textlength(text, font=current_font) <= available:
        return text
    suffix = "…"
    while text and draw.textlength(text + suffix, font=current_font) > available:
        text = text[:-1]
    return text + suffix


def detail_bounds(model: SheetModel) -> tuple[int, int, int, int]:
    return (
        model.first_row,
        min(model.last_row, model.first_row + DETAIL_ROWS - 1),
        model.first_column,
        min(model.last_column, model.first_column + DETAIL_COLUMNS - 1),
    )


def render_preview(model: SheetModel, source_name: str, size: tuple[int, int]) -> Image.Image:
    width, height = size
    image = Image.new("RGB", size, "#f7fafc")
    draw = ImageDraw.Draw(image)
    title_font = font(max(15, width // 44), bold=True)
    meta_font = font(max(10, width // 78))
    cell_fonts: dict[tuple[int, bool], ImageFont.FreeTypeFont] = {}

    def cell_font(size: int, bold: bool) -> ImageFont.FreeTypeFont:
        key = (size, bold)
        if key not in cell_fonts:
            cell_fonts[key] = font(size, bold=bold)
        return cell_fonts[key]
    padding = max(16, width // 45)
    header_height = max(72, height // 9)
    draw.rounded_rectangle((padding, padding, width - padding, height - padding), radius=14, fill="#ffffff", outline="#cbd8e0", width=2)
    draw.text((padding + 18, padding + 14), trim_text(draw, source_name, width - padding * 2 - 36, title_font), font=title_font, fill="#1d2b36")
    meta = f"Лист: {model.name}  •  Диапазон: {model.range_label}  •  Заполнено ячеек: {model.nonempty_cells}"
    draw.text((padding + 18, padding + 18 + title_font.size + 4), trim_text(draw, meta, width - padding * 2 - 36, meta_font), font=meta_font, fill="#617482")

    first_row, last_row, first_column, last_column = detail_bounds(model)
    content_left = padding + 18
    content_top = padding + header_height
    content_right = width - padding - 18
    content_bottom = height - padding - 18
    available_width = content_right - content_left
    available_height = content_bottom - content_top
    raw_column_widths = [max(38.0, min(260.0, model.column_widths[column] * 7.0)) for column in range(first_column, last_column + 1)]
    raw_row_heights = [max(18.0, min(80.0, model.row_heights[row] * 1.3)) for row in range(first_row, last_row + 1)]
    scale = min(available_width / max(1, sum(raw_column_widths)), available_height / max(1, sum(raw_row_heights)), 1.0)
    column_widths = {column: raw_column_widths[column - first_column] * scale for column in range(first_column, last_column + 1)}
    row_heights = {row: raw_row_heights[row - first_row] * scale for row in range(first_row, last_row + 1)}
    x_positions = {first_column: content_left}
    y_positions = {first_row: content_top}
    for column in range(first_column + 1, last_column + 1):
        x_positions[column] = x_positions[column - 1] + column_widths[column - 1]
    for row in range(first_row + 1, last_row + 1):
        y_positions[row] = y_positions[row - 1] + row_heights[row - 1]

    draw.rectangle((content_left, content_top, content_right, content_bottom), fill="#ffffff", outline="#aebec8", width=1)
    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            if (row, column) in model.merged_children:
                continue
            cell = model.cells.get((row, column))
            x1, y1 = x_positions[column], y_positions[row]
            span_rows, span_columns = model.merged_starts.get((row, column), (1, 1))
            end_column = min(last_column, column + span_columns - 1)
            end_row = min(last_row, row + span_rows - 1)
            x2 = x_positions[end_column] + column_widths[end_column]
            y2 = y_positions[end_row] + row_heights[end_row]
            draw.rectangle((x1, y1, x2, y2), fill=cell.fill if cell and cell.fill else "#ffffff", outline="#c8d2d8", width=1)
            if not cell:
                continue
            # Excel stores font sizes in points, rows in points, and a Windows
            # pixel is roughly 96/72 of a point.  The old renderer instead
            # used the *image width* to choose a font size: a high-resolution
            # card therefore made 31 px letters inside 19 px rows.  The font
            # now comes from the source cell and can never exceed its box.
            authored_px = cell.font_size * (96 / 72) * scale
            font_size = max(6, min(round(authored_px), round((y2 - y1) * 0.85)))
            current_font = cell_font(font_size, cell.bold)
            # Font size is not ink height: Segoe UI has a non-zero baseline
            # offset and descender.  Reduce until its actual metrics fit,
            # otherwise text sits on the bottom grid line and looks clipped.
            while font_size > 6 and sum(current_font.getmetrics()) > (y2 - y1) * 0.85:
                font_size -= 1
                current_font = cell_font(font_size, cell.bold)
            inset = max(2, round(4 * scale))
            text = trim_text(draw, cell.value, max(4, int(x2 - x1 - inset * 2)), current_font)
            if cell.horizontal == "center":
                text_width = draw.textlength(text, font=current_font)
                text_x = x1 + max(inset, (x2 - x1 - text_width) / 2)
            elif cell.horizontal == "right":
                text_width = draw.textlength(text, font=current_font)
                text_x = max(x1 + inset, x2 - inset - text_width)
            else:
                text_x = x1 + inset
            bbox = current_font.getbbox(text)
            ink_height = bbox[3] - bbox[1]
            text_y = y1 + max(1, (y2 - y1 - ink_height) / 2) - bbox[1]
            draw.text((text_x, text_y), text, font=current_font, fill=cell.font_color or "#24333d")
    return image


def render(path: Path, output_root: Path, preview_size: tuple[int, int] = PREVIEW_SIZE) -> dict[str, Any]:
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("Первый прототип принимает XLSX/XLSM. XLS будет отдельным сравнительным тестом.")
    started = time.perf_counter()
    key = source_key(path, preview_size)
    cache_dir = output_root / key
    manifest_path = cache_dir / "manifest.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest["cacheHit"] = True
        manifest["elapsedMs"] = round((time.perf_counter() - started) * 1_000)
        return manifest
    cache_dir.mkdir(parents=True, exist_ok=True)
    styles_book = openpyxl.load_workbook(path, read_only=False, data_only=False)
    values_book = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        model = select_primary_sheet(styles_book, values_book)
    finally:
        styles_book.close()
        values_book.close()
    preview = render_preview(model, path.name, preview_size)
    thumbnail = preview.copy()
    thumbnail.thumbnail(THUMBNAIL_SIZE, Image.Resampling.LANCZOS)
    preview_path = cache_dir / "preview.png"
    thumbnail_path = cache_dir / "thumbnail.webp"
    preview.save(preview_path, "PNG", optimize=True)
    thumbnail.save(thumbnail_path, "WEBP", quality=86, method=6)
    manifest = {
        "rendererVersion": RENDERER_VERSION,
        "previewPixels": {"width": preview_size[0], "height": preview_size[1]},
        "cacheHit": False,
        "source": str(path.resolve()),
        "cacheKey": key,
        "sheet": {"index": model.index, "name": model.name, "range": model.range_label, "nonemptyCells": model.nonempty_cells},
        "preview": str(preview_path.resolve()),
        "thumbnail": str(thumbnail_path.resolve()),
        "elapsedMs": round((time.perf_counter() - started) * 1_000),
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="F-Engineering Excel fast preview prototype")
    parser.add_argument("input", type=Path, help="Путь к XLSX/XLSM")
    parser.add_argument("--output", type=Path, default=Path("runtime/cache/excel-fast-preview"), help="Папка cache")
    parser.add_argument("--size", default="1400x900", help="Размер крупной карточки, например 3600x2300")
    args = parser.parse_args()
    if not args.input.exists() or not args.input.is_file():
        raise SystemExit(f"Файл не найден: {args.input}")
    try:
        width, height = (int(value) for value in args.size.lower().split("x", maxsplit=1))
    except ValueError as error:
        raise SystemExit("Размер должен быть в формате ШИРИНАxВЫСОТА, например 3600x2300") from error
    if not (320 <= width <= 6000 and 240 <= height <= 6000):
        raise SystemExit("Размер preview должен быть в диапазоне 320–6000 px по каждой стороне")
    print(json.dumps(render(args.input, args.output, (width, height)), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
