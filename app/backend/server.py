from __future__ import annotations

import argparse
import hashlib
import html
import json
import mimetypes
import os
import shutil
import subprocess
import time
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import unquote, urlparse

try:
    import openpyxl
except ImportError:  # pragma: no cover - reported through the local API
    openpyxl = None


REPO_ROOT = Path(__file__).resolve().parents[2]
FRONTEND_DIR = REPO_ROOT / "app" / "frontend"
RUNTIME_DIR = REPO_ROOT / "runtime"
MANIFESTS_DIR = RUNTIME_DIR / "manifests"
PDF_CACHE_DIR = RUNTIME_DIR / "cache" / "pdf"
WORD_CACHE_DIR = RUNTIME_DIR / "cache" / "word"
EXCEL_CACHE_DIR = RUNTIME_DIR / "cache" / "excel"
DWG_CACHE_DIR = RUNTIME_DIR / "cache" / "dwg"
WORD_CONVERT_SCRIPT = REPO_ROOT / "scripts" / "convert_word_to_pdf.ps1"
EXCEL_CONVERT_SCRIPT = REPO_ROOT / "scripts" / "convert_excel_to_pdf.ps1"
DWG_RENDER_SCRIPT = REPO_ROOT / "scripts" / "render_dwg_model_space.ps1"
DWG_REVIEW_OPEN_SCRIPT = REPO_ROOT / "scripts" / "open_dwg_review_copy.ps1"
VERSION = "0.4.0-v3-pdf-render"
SKIP_DIR_NAMES = {".git", "__pycache__", "node_modules", ".venv", "venv"}
DEFAULT_PDF_DPI = 300
PDF_PAGE_TIMEOUT_SECONDS = 25
PDF_DOCUMENT_TIMEOUT_SECONDS = 75
WORD_CONVERT_TIMEOUT_SECONDS = 120
EXCEL_CONVERT_TIMEOUT_SECONDS = 180
DWG_RENDER_TIMEOUT_SECONDS = 180
DWG_MODEL_PAGE_TIMEOUT_SECONDS = 75
DWG_OPEN_TIMEOUT_SECONDS = 60
MAX_XLSX_ROWS = 2000
MAX_XLSX_COLS = 100
POPPLER_BIN_DIR = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "native"
    / "poppler"
    / "Library"
    / "bin"
)


def object_id_for_path(path: Path) -> str:
    return hashlib.sha1(str(path.resolve()).casefold().encode("utf-8")).hexdigest()[:16]


def file_extension(path: Path) -> str:
    suffix = path.suffix.upper().lstrip(".")
    return suffix or "NO_EXT"


def build_tree(folder: Path) -> tuple[dict, dict[str, int], int, int]:
    counts: dict[str, int] = {}
    folder_count = 0
    file_count = 0

    def walk(current: Path) -> dict:
        nonlocal folder_count, file_count
        folder_count += 1
        children = []
        try:
            entries = sorted(current.iterdir(), key=lambda item: (not item.is_dir(), item.name.casefold()))
        except OSError as error:
            return {
                "type": "folder",
                "name": current.name,
                "path": str(current),
                "error": str(error),
                "children": [],
            }

        for entry in entries:
            if entry.is_dir():
                if entry.name in SKIP_DIR_NAMES:
                    continue
                children.append(walk(entry))
            elif entry.is_file():
                ext = file_extension(entry)
                counts[ext] = counts.get(ext, 0) + 1
                file_count += 1
                children.append(
                    {
                        "type": "file",
                        "name": entry.name,
                        "path": str(entry),
                        "extension": ext,
                        "size": entry.stat().st_size,
                    }
                )

        return {"type": "folder", "name": current.name, "path": str(current), "children": children}

    return walk(folder), counts, folder_count, file_count


def scan_object(raw_path: str) -> dict:
    if not raw_path or not raw_path.strip():
        raise ValueError("Путь к папке объекта пустой")
    root = Path(raw_path.strip().strip('"')).expanduser()
    if not root.exists():
        raise FileNotFoundError(f"Папка не найдена: {root}")
    if not root.is_dir():
        raise NotADirectoryError(f"Это не папка: {root}")

    tree, extension_counts, folder_count, file_count = build_tree(root)
    manifest = {
        "id": object_id_for_path(root),
        "name": root.name,
        "rootPath": str(root.resolve()),
        "scannedAt": datetime.now().isoformat(timespec="seconds"),
        "statistics": {
            "folders": folder_count,
            "files": file_count,
            "extensions": extension_counts,
        },
        "tree": tree,
    }
    MANIFESTS_DIR.mkdir(parents=True, exist_ok=True)
    (MANIFESTS_DIR / f"{manifest['id']}.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return manifest


def manifest_path(object_id: str) -> Path:
    return MANIFESTS_DIR / f"{object_id}.json"


def load_manifest(object_id: str) -> dict | None:
    path = manifest_path(object_id)
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def manifest_summary(manifest: dict) -> dict:
    return {
        "id": manifest.get("id"),
        "name": manifest.get("name"),
        "rootPath": manifest.get("rootPath"),
        "scannedAt": manifest.get("scannedAt"),
        "statistics": manifest.get("statistics", {}),
    }


def list_object_summaries() -> list[dict]:
    MANIFESTS_DIR.mkdir(parents=True, exist_ok=True)
    manifests: list[dict] = []
    for path in MANIFESTS_DIR.glob("*.json"):
        try:
            manifest = json.loads(path.read_text(encoding="utf-8"))
            manifests.append(manifest_summary(manifest))
        except (OSError, json.JSONDecodeError):
            continue
    return sorted(manifests, key=lambda item: str(item.get("scannedAt", "")), reverse=True)


def pdf_cache_key(path: Path, dpi: int) -> str:
    stat = path.stat()
    raw = f"{path.resolve()}|{stat.st_mtime_ns}|{stat.st_size}|{dpi}"
    return hashlib.sha1(raw.casefold().encode("utf-8")).hexdigest()[:20]


def file_cache_key(path: Path, purpose: str) -> str:
    stat = path.stat()
    raw = f"{purpose}|{path.resolve()}|{stat.st_mtime_ns}|{stat.st_size}"
    return hashlib.sha1(raw.casefold().encode("utf-8")).hexdigest()[:20]


def is_word_file(path: Path) -> bool:
    return path.suffix.casefold() in {".doc", ".docx"}


def is_excel_file(path: Path) -> bool:
    return path.suffix.casefold() in {".xls", ".xlsx", ".xlsm"}


def pdf_page_item(path: Path, key: str, page: int, png: Path) -> dict:
    return {
        "page": page,
        "name": f"{path.name} · стр. {page}",
        "url": f"/cache/pdf/{key}/{png.name}",
        "bytes": png.stat().st_size,
    }


def read_pdf_cache_manifest(path: Path, dpi: int, key: str, target_dir: Path) -> dict | None:
    manifest_path = target_dir / "manifest.json"
    manifest = None
    if manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            manifest = None
        if manifest and (manifest.get("path") != str(path) or manifest.get("dpi") != dpi or manifest.get("cacheKey") != key):
            manifest = None
    if not manifest:
        legacy_items = []
        for png in sorted(target_dir.glob("page-*.png")):
            try:
                page = int(png.stem.removeprefix("page-"))
            except ValueError:
                continue
            if page > 0 and png.stat().st_size > 0:
                legacy_items.append(pdf_page_item(path, key, page, png))
        if not legacy_items:
            return None
        return {
            "name": path.name,
            "path": str(path),
            "dpi": dpi,
            "pages": max(item["page"] for item in legacy_items),
            "renderedPages": len(legacy_items),
            "cacheKey": key,
            "complete": False,
            "errors": [],
            "items": legacy_items,
        }
    items = []
    for item in manifest.get("items", []):
        page = int(item.get("page") or 0)
        png = target_dir / f"page-{page}.png"
        if page > 0 and png.exists() and png.stat().st_size > 0:
            items.append(pdf_page_item(path, key, page, png))
    if not items:
        return None
    manifest["items"] = items
    return manifest


def write_pdf_cache_manifest(
    path: Path,
    dpi: int,
    key: str,
    target_dir: Path,
    page_count: int,
    items: list[dict],
    errors: list[dict],
) -> None:
    manifest = {
        "name": path.name,
        "path": str(path),
        "dpi": dpi,
        "pages": page_count,
        "renderedPages": len(items),
        "cacheKey": key,
        "complete": len(items) == page_count and not errors,
        "errors": errors,
        "items": items,
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
    }
    (target_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def poppler_tool(name: str) -> str | None:
    exe = POPPLER_BIN_DIR / f"{name}.exe"
    if exe.exists():
        return str(exe)
    found = shutil.which(name)
    if found and not found.lower().endswith(".cmd"):
        return found
    return found


def run_poppler(args: list[str], timeout: int) -> subprocess.CompletedProcess[str]:
    process = subprocess.Popen(
        args,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    try:
        stdout, stderr = process.communicate(timeout=timeout)
    except subprocess.TimeoutExpired as error:
        process.kill()
        stdout, stderr = process.communicate()
        raise subprocess.TimeoutExpired(args, timeout, output=stdout, stderr=stderr) from error
    return subprocess.CompletedProcess(args=args, returncode=process.returncode, stdout=stdout, stderr=stderr)


def pdf_page_count(path: Path) -> int:
    pdfinfo = poppler_tool("pdfinfo")
    if not pdfinfo:
        raise RuntimeError("pdfinfo не найден. Нужен Poppler из runtime.")
    result = run_poppler(
        [pdfinfo, str(path)],
        timeout=60,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "pdfinfo не смог прочитать PDF")
    for line in result.stdout.splitlines():
        if line.startswith("Pages:"):
            return int(line.split(":", 1)[1].strip())
    raise RuntimeError("pdfinfo не вернул количество страниц PDF")


def word_to_pdf(path: Path) -> tuple[Path, bool]:
    if not path.exists():
        raise FileNotFoundError(f"Word-файл не найден: {path}")
    if not path.is_file() or not is_word_file(path):
        raise ValueError(f"Это не Word-файл: {path}")
    if not WORD_CONVERT_SCRIPT.exists():
        raise RuntimeError(f"Скрипт конвертации Word не найден: {WORD_CONVERT_SCRIPT}")

    key = file_cache_key(path, "word-pdf")
    target_dir = WORD_CACHE_DIR / key
    target_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = target_dir / f"{path.stem}.pdf"
    manifest_path = target_dir / "manifest.json"

    cached = False
    if pdf_path.exists() and pdf_path.stat().st_size > 0 and manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            cached = (
                manifest.get("sourcePath") == str(path)
                and manifest.get("cacheKey") == key
                and manifest.get("sourceMtimeNs") == path.stat().st_mtime_ns
                and manifest.get("sourceSize") == path.stat().st_size
            )
        except (OSError, json.JSONDecodeError):
            cached = False

    if cached:
        return pdf_path, True

    if pdf_path.exists():
        pdf_path.unlink()

    process = subprocess.run(
        [
            "powershell",
            "-STA",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(WORD_CONVERT_SCRIPT),
            "-InputPath",
            str(path),
            "-OutputPath",
            str(pdf_path),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=WORD_CONVERT_TIMEOUT_SECONDS,
    )
    if process.returncode != 0:
        message = process.stderr.strip() or process.stdout.strip() or "Word не смог конвертировать документ в PDF"
        raise RuntimeError(message)
    if not pdf_path.exists() or pdf_path.stat().st_size <= 0:
        raise RuntimeError("Word не создал PDF для preview")

    manifest_path.write_text(
        json.dumps(
            {
                "sourcePath": str(path),
                "sourceName": path.name,
                "sourceMtimeNs": path.stat().st_mtime_ns,
                "sourceSize": path.stat().st_size,
                "cacheKey": key,
                "pdfPath": str(pdf_path),
                "convertedAt": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return pdf_path, False


def render_word(path: Path, dpi: int = DEFAULT_PDF_DPI) -> dict:
    pdf_path, convert_cache_hit = word_to_pdf(path)
    document = render_pdf(pdf_path, dpi=dpi)
    document["sourcePath"] = str(path)
    document["sourceName"] = path.name
    document["sourceType"] = file_extension(path)
    document["convertedPdfPath"] = str(pdf_path)
    document["convertCacheHit"] = convert_cache_hit
    return document


def dwg_to_model_pdf(path: Path) -> tuple[Path, bool]:
    """Create a cached read-only Model Space overview through installed ZWCAD.

    This intentionally produces an overview, not editable CAD geometry and not
    a substitute for sheets/layouts.  The source DWG is only opened read-only.
    """
    if not path.exists():
        raise FileNotFoundError(f"DWG-файл не найден: {path}")
    if not path.is_file() or path.suffix.casefold() != ".dwg":
        raise ValueError(f"Это не DWG-файл: {path}")
    if not DWG_RENDER_SCRIPT.exists():
        raise RuntimeError(f"Скрипт рендера DWG не найден: {DWG_RENDER_SCRIPT}")

    key = file_cache_key(path, "dwg-model-a0-v1")
    target_dir = DWG_CACHE_DIR / key
    target_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = target_dir / "model-space-a0.pdf"
    manifest_path = target_dir / "manifest.json"
    if pdf_path.exists() and pdf_path.stat().st_size > 1024 and manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            if (
                manifest.get("sourcePath") == str(path)
                and manifest.get("cacheKey") == key
                and manifest.get("sourceMtimeNs") == path.stat().st_mtime_ns
                and manifest.get("sourceSize") == path.stat().st_size
            ):
                return pdf_path, True
        except (OSError, json.JSONDecodeError):
            pass

    if pdf_path.exists():
        pdf_path.unlink()
    process = subprocess.run(
        [
            "powershell",
            "-STA",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(DWG_RENDER_SCRIPT),
            "-InputPath",
            str(path),
            "-OutputPath",
            str(pdf_path),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=DWG_RENDER_TIMEOUT_SECONDS,
    )
    if process.returncode != 0:
        message = process.stderr.strip() or process.stdout.strip() or "ZWCAD не смог создать Model Space preview"
        raise RuntimeError(message)
    if not pdf_path.exists() or pdf_path.stat().st_size <= 1024:
        raise RuntimeError("ZWCAD не создал PDF Model Space для preview")

    manifest_path.write_text(
        json.dumps(
            {
                "sourcePath": str(path),
                "sourceName": path.name,
                "sourceMtimeNs": path.stat().st_mtime_ns,
                "sourceSize": path.stat().st_size,
                "cacheKey": key,
                "pdfPath": str(pdf_path),
                "mode": "model-space-a0-extents-no-lineweights",
                "renderedAt": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return pdf_path, False


def render_dwg_model(path: Path, dpi: int = DEFAULT_PDF_DPI) -> dict:
    pdf_path, convert_cache_hit = dwg_to_model_pdf(path)
    document = render_pdf(pdf_path, dpi=dpi, page_timeout_seconds=DWG_MODEL_PAGE_TIMEOUT_SECONDS)
    document["name"] = path.name
    document["sourcePath"] = str(path)
    document["sourceName"] = path.name
    document["sourceType"] = "DWG"
    document["convertedPdfPath"] = str(pdf_path)
    document["convertCacheHit"] = convert_cache_hit
    document["previewMode"] = "model-space-a0"
    return document


def open_dwg_for_review(path: Path) -> Path:
    """Open the original DWG through the registered Windows application.

    ZWCAD COM automation creates hidden ``/Automation`` instances which can
    hang while a drawing is opening from Google Drive.  ShellExecute follows
    the same association and DDE path as a user double-click in Explorer.
    """
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"DWG-файл не найден: {path}")
    if os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
    else:
        subprocess.Popen(["xdg-open", str(path)])
    return path


def append_native_open_log(event: dict) -> None:
    """Write one UTF-8 diagnostic record for a native-open attempt."""
    logs_dir = RUNTIME_DIR / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    with (logs_dir / "native-open.jsonl").open("a", encoding="utf-8") as log:
        log.write(json.dumps(event, ensure_ascii=False) + "\n")


def excel_html_cache_dir(path: Path) -> Path:
    """Return an immutable cache location for one exact workbook revision."""
    return EXCEL_CACHE_DIR / "html" / file_cache_key(path, "excel-html")


def excel_cell_color(color: object) -> str | None:
    value = getattr(color, "rgb", None)
    value = str(value) if value is not None else ""
    if not value or len(value) < 6 or (len(value) == 8 and value[:2] == "00"):
        return None
    return f"#{value[-6:]}"


def excel_sheet_html(path: Path, sheet_index: int) -> tuple[str, dict]:
    """Create a read-only HTML sheet with authored geometry.

    The Launcher is a fast navigator, not a replacement for Excel.  The limit is
    intentional: long operational registers are opened in the native program.
    """
    if openpyxl is None:
        raise RuntimeError("Для HTML-просмотра Excel нужен пакет openpyxl")
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("HTML-просмотр пока поддерживает XLSX/XLSM. Откройте XLS в Excel.")

    styles_book = openpyxl.load_workbook(path, read_only=False, data_only=False)
    values_book = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = styles_book.worksheets[sheet_index]
        values = values_book.worksheets[sheet_index]
        # Some valid workbooks (notably registers exported by third-party
        # systems) have no stored dimension in one of the loaded views.
        # openpyxl then returns None, which must mean an empty 1x1 sheet here,
        # not an unhandled server exception and a blank Launcher screen.
        sheet_max_row = int(sheet.max_row or 1)
        values_max_row = int(values.max_row or 1)
        sheet_max_column = int(sheet.max_column or 1)
        values_max_column = int(values.max_column or 1)
        max_column = min(max(sheet_max_column, values_max_column), MAX_XLSX_COLS)
        columns = [
            column
            for column in range(1, max_column + 1)
            if not sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].hidden
        ]
        column_pixels = {}
        for column in columns:
            width = sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width or 8.43
            column_pixels[column] = max(4, min(720, round(width * 7 + 5)))

        merged_start = {}
        merged_skip = set()
        for area in sheet.merged_cells.ranges:
            shown_columns = [column for column in columns if area.min_col <= column <= area.max_col]
            if not shown_columns:
                continue
            start = (area.min_row, shown_columns[0])
            merged_start[start] = (len(shown_columns), area.max_row - area.min_row + 1)
            for row in range(area.min_row, area.max_row + 1):
                for column in shown_columns:
                    if (row, column) != start:
                        merged_skip.add((row, column))

        first_value_row = 0
        last_value_row = 0
        for row in range(1, min(values_max_row, MAX_XLSX_ROWS) + 1):
            if any(values.cell(row, column).value is not None for column in columns):
                if not first_value_row:
                    first_value_row = row
                last_value_row = row
        last_merge_row = max((area.max_row for area in sheet.merged_cells.ranges), default=0)
        rendered_rows = min(max(last_value_row, last_merge_row), MAX_XLSX_ROWS)
        was_limited = max(sheet_max_row, values_max_row) > MAX_XLSX_ROWS

        first_rendered_row = first_value_row or 1
        rows = []
        for row in range(first_rendered_row, rendered_rows + 1):
            if sheet.row_dimensions[row].hidden:
                continue
            cells = []
            for column in columns:
                if (row, column) in merged_skip:
                    continue
                cell = sheet.cell(row, column)
                raw_value = values.cell(row, column).value
                value = "" if raw_value is None else str(raw_value)
                css = []
                # A single neutral grid is supplied by the page stylesheet.
                # Re-emitting four border declarations for every cell made a
                # 2,000-row workbook produce 8–9 MB of HTML and a long white
                # screen before Chrome could paint it.  This is a navigator,
                # so fast first paint has priority over reproducing every
                # individual spreadsheet border colour.
                if cell.font.bold:
                    css.append("font-weight:500")
                if cell.font.italic:
                    css.append("font-style:italic")
                # Defaults are already declared once in the HTML stylesheet.
                # Repeating Calibri 11px on every one of 52,000 cells turns a
                # normal workbook into a multi-megabyte page that opens white.
                if cell.font.sz and round(cell.font.sz) != 11:
                    css.append(f"font-size:{max(6, min(32, cell.font.sz))}px")
                if cell.font.name and cell.font.name.casefold() not in {"calibri", "arial", "segoe ui"}:
                    css.append(f"font-family:{html.escape(cell.font.name, quote=True)}")
                font_color = excel_cell_color(cell.font.color)
                if font_color:
                    css.append(f"color:{font_color}")
                fill = excel_cell_color(cell.fill.fgColor)
                if cell.fill.fill_type == "solid" and fill:
                    css.append(f"background:{fill}")
                if cell.alignment.horizontal in {"left", "center", "right"}:
                    css.append(f"text-align:{cell.alignment.horizontal}")
                if cell.alignment.vertical in {"top", "center", "bottom"}:
                    css.append(f"vertical-align:{cell.alignment.vertical}")
                if cell.alignment.wrap_text is False:
                    css.append("white-space:pre;overflow:hidden")
                colspan, rowspan = merged_start.get((row, column), (1, 1))
                span = (f' colspan="{colspan}"' if colspan > 1 else "") + (f' rowspan="{rowspan}"' if rowspan > 1 else "")
                cells.append(f'<td{span} style="{";".join(css)}">{html.escape(value)}</td>')
            authored_height = sheet.row_dimensions[row].height
            row_style = f' style="height:{max(1, round(authored_height * 1.33))}px"' if authored_height else ""
            rows.append(f"<tr{row_style}><th>{row}</th>{''.join(cells)}</tr>")

        cols = '<col class="row-number">' + "".join(
            f'<col style="width:{column_pixels[column]}px">' for column in columns
        )
        table_width = 38 + sum(column_pixels.values())
        notice = (
            f"Показаны первые {MAX_XLSX_ROWS:,} строк. Полный рабочий файл откройте в Excel."
            if was_limited else "Просмотр без редактирования"
        )
        rendered = {
            "name": sheet.title,
            "rows": max(0, rendered_rows - first_rendered_row + 1),
            "firstRow": first_rendered_row,
            "columns": len(columns),
            "limited": was_limited,
        }
        page = f'''<!doctype html><html lang="ru"><head><meta charset="utf-8"><title>{html.escape(sheet.title)}</title>
<style>
html,body{{margin:0;min-width:max-content;background:#fff;color:#20262d;font:11px "Segoe UI",Arial,sans-serif;overflow:auto}}
#sheet-canvas{{position:relative;transform-origin:0 0}}#sheet{{position:absolute;left:0;top:0;transform-origin:0 0}}
.notice{{position:sticky;top:0;z-index:2;padding:6px 10px;border-bottom:1px solid #d3dde4;background:#f7fafc;color:#607080;font-size:11px}}
table{{border-collapse:collapse;table-layout:fixed;width:{table_width}px}}col.row-number{{width:38px}}
th,td{{box-sizing:border-box;border:1px solid #cbd5dc;padding:2px 4px;vertical-align:top;white-space:pre-wrap;overflow-wrap:break-word}}
th{{position:sticky;left:0;z-index:1;background:#f1f5f7;color:#657687;font:10px "Segoe UI",Arial,sans-serif;text-align:right}}td{{overflow:hidden}}
</style></head><body><div id="sheet-canvas"><div id="sheet"><table><colgroup>{cols}</colgroup><tbody>{''.join(rows)}</tbody></table></div></div>
<script>const canvas=document.getElementById('sheet-canvas'),sheet=document.getElementById('sheet');let width=0,height=0,padding=0,scale=1,rotation=0,hand=true,dragging=false,startX=0,startY=0,startLeft=0,startTop=0;function dimensions(){{return Math.abs(rotation%180)===90?{{width:height,height:width}}:{{width,height}}}}function zoom(value){{if(!width){{width=sheet.offsetWidth;height=sheet.offsetHeight}}scale=Math.max(.35,Math.min(3,value));padding=Math.max(innerWidth,innerHeight);const size=dimensions();canvas.style.width=(size.width*scale+padding*2)+'px';canvas.style.height=(size.height*scale+padding*2)+'px';sheet.style.transform='translate('+padding+'px,'+padding+'px) rotate('+rotation+'deg) scale('+scale+')'}}function fit(){{if(!width)zoom(1);const size=dimensions(),value=Math.min(1,(innerWidth-48)/size.width,(innerHeight-48)/size.height);zoom(value);requestAnimationFrame(()=>{{scrollTo(padding,padding);parent.postMessage({{type:'launcher-sheet-fitted',value:scale}},'*')}})}}function cursor(){{document.body.style.cursor=hand?(dragging?'grabbing':'grab'):'default'}}addEventListener('load',()=>{{zoom(1);cursor();requestAnimationFrame(()=>scrollTo(padding,padding))}});addEventListener('wheel',event=>{{if(!event.ctrlKey)return;event.preventDefault();zoom(scale*(event.deltaY<0?1.12:.89))}},{{passive:false}});addEventListener('pointerdown',event=>{{if(!hand||event.button!==0)return;dragging=true;startX=event.clientX;startY=event.clientY;startLeft=scrollX;startTop=scrollY;document.body.setPointerCapture?.(event.pointerId);cursor();event.preventDefault()}});addEventListener('pointermove',event=>{{if(!dragging)return;scrollTo(startLeft-(event.clientX-startX),startTop-(event.clientY-startY))}});addEventListener('pointerup',event=>{{if(!dragging)return;dragging=false;document.body.releasePointerCapture?.(event.pointerId);cursor()}});addEventListener('pointercancel',()=>{{dragging=false;cursor()}});addEventListener('message',event=>{{if(!event.data)return;if(event.data.type==='launcher-sheet-zoom')zoom(event.data.value);if(event.data.type==='launcher-sheet-fit')fit();if(event.data.type==='launcher-sheet-rotate'){{rotation=((Number(event.data.value)||0)%360+360)%360;zoom(scale)}}if(event.data.type==='launcher-sheet-hand'){{hand=Boolean(event.data.value);dragging=false;cursor()}}}});</script></body></html>'''
        return page, rendered
    finally:
        styles_book.close()
        values_book.close()


def excel_sheet_preview(path: Path, sheet_index: int) -> dict:
    """Build one sheet on demand; opening a book must not wait for every tab."""
    cache_dir = excel_html_cache_dir(path)
    cache_dir.mkdir(parents=True, exist_ok=True)
    # A new cache version makes existing pages harmless without deleting a
    # user's cache, including the viewer controls embedded in this HTML.
    output = cache_dir / f"sheet-{sheet_index + 1}-v8.html"
    metadata_path = cache_dir / f"sheet-{sheet_index + 1}-v8.json"
    metadata = None
    if output.exists() and metadata_path.exists():
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            metadata = None
    if metadata is None:
        page, metadata = excel_sheet_html(path, sheet_index)
        output.write_text(page, encoding="utf-8")
        metadata_path.write_text(json.dumps(metadata, ensure_ascii=False), encoding="utf-8")
    return {
        "index": sheet_index,
        "url": f"/cache/excel/html/{cache_dir.name}/{output.name}",
        **metadata,
    }


def excel_thumbnail_preview(path: Path) -> str:
    """Build a visual card from the exact same HTML sheet as the large view.

    A workbook must have one visual source of truth.  The former thumbnail
    constructed a second, simplified table with its own font and row rules;
    it could therefore disagree with the readable sheet.  The card is now a
    scaled viewport of the cached HTML sheet used by the full viewer.
    """
    if openpyxl is None:
        raise RuntimeError("Для HTML-просмотра Excel нужен пакет openpyxl")
    # The rail already places this page in a scaled, clipped iframe.  Returning
    # the sheet directly means its geometry, fonts and initial position are
    # exactly the same in the card and in the large viewer.
    return excel_sheet_preview(path, 0)["url"]


def excel_workbook_preview(path: Path) -> dict:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Excel-файл не найден: {path}")
    if not is_excel_file(path):
        raise ValueError(f"Это не Excel-файл: {path}")
    if openpyxl is None:
        raise RuntimeError("Для HTML-просмотра Excel нужен пакет openpyxl")
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("HTML-просмотр пока поддерживает XLSX/XLSM. Откройте XLS в Excel.")

    book = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = book.sheetnames
    finally:
        book.close()

    sheets = [{"index": index, "name": name} for index, name in enumerate(sheet_names)]
    return {
        "name": path.name,
        "path": str(path),
        "sheets": sheets,
        "maxRows": MAX_XLSX_ROWS,
        "cacheKey": excel_html_cache_dir(path).name,
        "thumbnailUrl": excel_thumbnail_preview(path),
    }


def excel_to_pdf(path: Path) -> tuple[Path, bool]:
    if not path.exists():
        raise FileNotFoundError(f"Excel-файл не найден: {path}")
    if not path.is_file() or not is_excel_file(path):
        raise ValueError(f"Это не Excel-файл: {path}")
    if not EXCEL_CONVERT_SCRIPT.exists():
        raise RuntimeError(f"Скрипт конвертации Excel не найден: {EXCEL_CONVERT_SCRIPT}")

    key = file_cache_key(path, "excel-pdf")
    target_dir = EXCEL_CACHE_DIR / key
    target_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = target_dir / f"{path.stem}.pdf"
    manifest_path = target_dir / "manifest.json"

    cached = False
    if pdf_path.exists() and pdf_path.stat().st_size > 0 and manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            cached = (
                manifest.get("sourcePath") == str(path)
                and manifest.get("cacheKey") == key
                and manifest.get("sourceMtimeNs") == path.stat().st_mtime_ns
                and manifest.get("sourceSize") == path.stat().st_size
            )
        except (OSError, json.JSONDecodeError):
            cached = False

    if cached:
        return pdf_path, True

    if pdf_path.exists():
        pdf_path.unlink()

    process = subprocess.run(
        [
            "powershell",
            "-STA",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(EXCEL_CONVERT_SCRIPT),
            "-InputPath",
            str(path),
            "-OutputPath",
            str(pdf_path),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=EXCEL_CONVERT_TIMEOUT_SECONDS,
    )
    if process.returncode != 0:
        message = process.stderr.strip() or process.stdout.strip() or "Excel не смог экспортировать книгу в PDF"
        raise RuntimeError(message)
    if not pdf_path.exists() or pdf_path.stat().st_size <= 0:
        raise RuntimeError("Excel не создал PDF для preview")

    manifest_path.write_text(
        json.dumps(
            {
                "sourcePath": str(path),
                "sourceName": path.name,
                "sourceMtimeNs": path.stat().st_mtime_ns,
                "sourceSize": path.stat().st_size,
                "cacheKey": key,
                "pdfPath": str(pdf_path),
                "convertedAt": datetime.now().isoformat(timespec="seconds"),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return pdf_path, False


def render_excel(path: Path, dpi: int = DEFAULT_PDF_DPI) -> dict:
    pdf_path, convert_cache_hit = excel_to_pdf(path)
    document = render_pdf(pdf_path, dpi=dpi)
    document["sourcePath"] = str(path)
    document["sourceName"] = path.name
    document["sourceType"] = file_extension(path)
    document["convertedPdfPath"] = str(pdf_path)
    document["convertCacheHit"] = convert_cache_hit
    return document


def render_pdf(path: Path, dpi: int = DEFAULT_PDF_DPI, page_timeout_seconds: int = PDF_PAGE_TIMEOUT_SECONDS) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"PDF не найден: {path}")
    if not path.is_file() or path.suffix.casefold() != ".pdf":
        raise ValueError(f"Это не PDF-файл: {path}")

    key = pdf_cache_key(path, dpi)
    target_dir = PDF_CACHE_DIR / key
    target_dir.mkdir(parents=True, exist_ok=True)
    cached_manifest = read_pdf_cache_manifest(path, dpi, key, target_dir)
    if cached_manifest:
        cached_items = cached_manifest["items"]
        cached_errors = cached_manifest.get("errors", [])
        return {
            "name": path.name,
            "path": str(path),
            "dpi": dpi,
            "pages": cached_manifest.get("pages", len(cached_items)),
            "renderedPages": len(cached_items),
            "cacheKey": key,
            "cacheHit": True,
            "cacheHitPages": len(cached_items),
            "newRenderedPages": 0,
            "errors": cached_errors,
            "items": cached_items,
        }

    pdftoppm = poppler_tool("pdftoppm")
    if not pdftoppm:
        raise RuntimeError("pdftoppm не найден. Нужен Poppler из runtime.")

    page_count = pdf_page_count(path)
    started_at = time.monotonic()

    pages = []
    errors = []
    rendered_count = 0
    cache_hit_count = 0

    for page in range(1, page_count + 1):
        if time.monotonic() - started_at > PDF_DOCUMENT_TIMEOUT_SECONDS:
            for skipped_page in range(page, page_count + 1):
                errors.append({"page": skipped_page, "error": f"PDF остановлен по лимиту {PDF_DOCUMENT_TIMEOUT_SECONDS} сек. на файл."})
            break

        png = target_dir / f"page-{page}.png"
        if png.exists() and png.stat().st_size > 0:
            cache_hit_count += 1
        else:
            for stale in target_dir.glob(f"page-{page}*.png"):
                stale.unlink()
            prefix = target_dir / f"page-{page}"
            try:
                result = run_poppler(
                    [
                        pdftoppm,
                        "-f",
                        str(page),
                        "-l",
                        str(page),
                        "-singlefile",
                        "-r",
                        str(dpi),
                        "-png",
                        str(path),
                        str(prefix),
                    ],
                        timeout=page_timeout_seconds,
                )
                candidate = target_dir / f"page-{page}.png"
                if result.returncode != 0:
                    raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "pdftoppm не смог отрендерить страницу")
                if not candidate.exists() or candidate.stat().st_size <= 0:
                    raise RuntimeError("pdftoppm не создал PNG страницы")
                rendered_count += 1
            except subprocess.TimeoutExpired:
                errors.append({"page": page, "error": f"Таймаут рендера страницы {page}: {page_timeout_seconds} сек."})
                continue
            except (RuntimeError, OSError) as error:
                errors.append({"page": page, "error": str(error)})
                continue

        if png.exists() and png.stat().st_size > 0:
            pages.append(pdf_page_item(path, key, page, png))

    if not pages:
        result = run_poppler(
            [pdftoppm, "-v"],
            timeout=10,
        )
        tool_version = (result.stderr or result.stdout).strip()
        raise RuntimeError(f"Не удалось отрендерить ни одной страницы PDF. {tool_version}. Ошибки: {errors}")

    write_pdf_cache_manifest(path, dpi, key, target_dir, page_count, pages, errors)
    return {
        "name": path.name,
        "path": str(path),
        "dpi": dpi,
        "pages": page_count,
        "renderedPages": len(pages),
        "cacheKey": key,
        "cacheHit": cache_hit_count == page_count,
        "cacheHitPages": cache_hit_count,
        "newRenderedPages": rendered_count,
        "errors": errors,
        "items": pages,
    }


def render_pdf_page(path: Path, page: int, dpi: int = DEFAULT_PDF_DPI, page_timeout_seconds: int = PDF_PAGE_TIMEOUT_SECONDS) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"PDF не найден: {path}")
    if not path.is_file() or path.suffix.casefold() != ".pdf":
        raise ValueError(f"Это не PDF-файл: {path}")

    pdftoppm = poppler_tool("pdftoppm")
    if not pdftoppm:
        raise RuntimeError("pdftoppm не найден. Нужен Poppler из runtime.")

    page_count = pdf_page_count(path)
    if page < 1 or page > page_count:
        raise ValueError(f"Страница {page} вне диапазона 1-{page_count}")

    key = pdf_cache_key(path, dpi)
    target_dir = PDF_CACHE_DIR / key
    target_dir.mkdir(parents=True, exist_ok=True)
    png = target_dir / f"page-{page}.png"
    cache_hit = png.exists() and png.stat().st_size > 0

    if not cache_hit:
        for stale in target_dir.glob(f"page-{page}*.png"):
            stale.unlink()
        prefix = target_dir / f"page-{page}"
        try:
            result = run_poppler(
                [
                    pdftoppm,
                    "-f",
                    str(page),
                    "-l",
                    str(page),
                    "-singlefile",
                    "-r",
                    str(dpi),
                    "-png",
                    str(path),
                    str(prefix),
                ],
                timeout=page_timeout_seconds,
            )
        except subprocess.TimeoutExpired as error:
            raise RuntimeError(f"Таймаут рендера страницы {page}: {page_timeout_seconds} сек.") from error
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "pdftoppm не смог отрендерить страницу")
    if not png.exists() or png.stat().st_size <= 0:
        raise RuntimeError("pdftoppm не создал PNG страницы")

    item = pdf_page_item(path, key, page, png)
    cached_manifest = read_pdf_cache_manifest(path, dpi, key, target_dir)
    items_by_page = {existing["page"]: existing for existing in (cached_manifest or {}).get("items", [])}
    items_by_page[page] = item
    items = [items_by_page[index] for index in sorted(items_by_page)]
    write_pdf_cache_manifest(path, dpi, key, target_dir, page_count, items, (cached_manifest or {}).get("errors", []))

    return {
        "name": path.name,
        "path": str(path),
        "dpi": dpi,
        "pages": page_count,
        "page": page,
        "cacheKey": key,
        "cacheHit": cache_hit,
        "item": item,
    }


class LauncherHandler(BaseHTTPRequestHandler):
    server_version = "FEngineeringLauncherV3/0.1"

    def log_message(self, format: str, *args: object) -> None:
        logs_dir = RUNTIME_DIR / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        with (logs_dir / "server.log").open("a", encoding="utf-8") as log:
            log.write("%s - %s\n" % (self.log_date_time_string(), format % args))

    def send_json(self, status: HTTPStatus, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self.send_json(
                HTTPStatus.OK,
                {
                    "status": "ok",
                    "service": "F-Engineering Launcher v3",
                    "version": VERSION,
                    "repoRoot": str(REPO_ROOT),
                    "runtime": str(RUNTIME_DIR),
                    "wordPreview": True,
                    "excelPreview": True,
                },
            )
            return

        if parsed.path == "/api/launcher-info":
            self.send_json(
                HTTPStatus.OK,
                {
                    "version": VERSION,
                    "frontend": str(FRONTEND_DIR),
                    "runtime": str(RUNTIME_DIR),
                    "port": self.server.server_port,
                },
            )
            return

        if parsed.path == "/api/objects":
            self.send_json(HTTPStatus.OK, {"items": list_object_summaries()})
            return

        if parsed.path.startswith("/api/objects/"):
            object_id = unquote(parsed.path.removeprefix("/api/objects/")).strip()
            manifest = load_manifest(object_id)
            if manifest:
                self.send_json(HTTPStatus.OK, manifest)
            else:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "Объект не найден в manifest-хранилище"})
            return

        if parsed.path.startswith("/cache/"):
            self.serve_cache(parsed.path)
            return

        self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/objects/import":
            try:
                body = self.read_json()
                manifest = scan_object(str(body.get("path", "")))
                self.send_json(HTTPStatus.OK, manifest)
            except (ValueError, FileNotFoundError, NotADirectoryError, OSError, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/objects/exclude":
            try:
                body = self.read_json()
                object_id = str(body.get("id", "")).strip()
                if not object_id:
                    raise ValueError("Не выбран объект для исключения")
                target = MANIFESTS_DIR / f"{object_id}.json"
                if target.exists():
                    target.unlink()
                self.send_json(HTTPStatus.OK, {"ok": True, "removed": object_id})
            except (ValueError, OSError, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/choose-folder":
            try:
                import tkinter as tk
                from tkinter import filedialog

                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                selected = filedialog.askdirectory(title="Выберите папку объекта для F-Engineering Launcher v3")
                root.destroy()
                self.send_json(HTTPStatus.OK, {"path": selected})
            except Exception as error:
                self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": f"Диалог выбора папки недоступен: {error}"})
            return

        if parsed.path == "/api/pdf/render":
            try:
                body = self.read_json()
                raw_files = body.get("files", [])
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not isinstance(raw_files, list) or not raw_files:
                    raise ValueError("Не выбраны PDF-файлы для отображения")
                if len(raw_files) > 25:
                    raise ValueError("За один раз пока можно отрендерить не больше 25 PDF")
                documents = [render_pdf(Path(str(file_path)), dpi=dpi) for file_path in raw_files]
                document_errors = [
                    {"document": document["name"], "path": document["path"], **error}
                    for document in documents
                    for error in document.get("errors", [])
                ]
                self.send_json(
                    HTTPStatus.OK,
                    {
                        "dpi": dpi,
                        "documents": documents,
                        "totalPages": sum(document["pages"] for document in documents),
                        "renderedPages": sum(document["renderedPages"] for document in documents),
                        "errors": document_errors,
                        "renderedAt": datetime.now().isoformat(timespec="seconds"),
                    },
                )
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/pdf/page":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                page = int(body.get("page") or 1)
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not raw_file:
                    raise ValueError("Не выбран PDF-файл для отображения")
                self.send_json(HTTPStatus.OK, render_pdf_page(Path(raw_file), page=page, dpi=dpi))
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/word/render":
            try:
                body = self.read_json()
                raw_files = body.get("files", [])
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not isinstance(raw_files, list) or not raw_files:
                    raise ValueError("Не выбраны Word-файлы для отображения")
                if len(raw_files) > 10:
                    raise ValueError("За один раз пока можно отрендерить не больше 10 Word-файлов")
                documents = [render_word(Path(str(file_path)), dpi=dpi) for file_path in raw_files]
                document_errors = [
                    {"document": document["name"], "path": document["path"], **error}
                    for document in documents
                    for error in document.get("errors", [])
                ]
                self.send_json(
                    HTTPStatus.OK,
                    {
                        "dpi": dpi,
                        "documents": documents,
                        "totalPages": sum(document["pages"] for document in documents),
                        "renderedPages": sum(document["renderedPages"] for document in documents),
                        "errors": document_errors,
                        "renderedAt": datetime.now().isoformat(timespec="seconds"),
                    },
                )
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/word/page":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                page = int(body.get("page") or 1)
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not raw_file:
                    raise ValueError("Не выбран Word-файл для отображения")
                pdf_path, convert_cache_hit = word_to_pdf(Path(raw_file))
                payload = render_pdf_page(pdf_path, page=page, dpi=dpi)
                payload["sourcePath"] = raw_file
                payload["sourceType"] = file_extension(Path(raw_file))
                payload["convertedPdfPath"] = str(pdf_path)
                payload["convertCacheHit"] = convert_cache_hit
                self.send_json(HTTPStatus.OK, payload)
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/dwg/model-render":
            try:
                body = self.read_json()
                raw_files = body.get("files", [])
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not isinstance(raw_files, list) or not raw_files:
                    raise ValueError("Не выбраны DWG-файлы для отображения")
                if len(raw_files) > 1:
                    raise ValueError("Model Space preview пока создаётся по одному DWG-файлу")
                documents = [render_dwg_model(Path(str(file_path)), dpi=dpi) for file_path in raw_files]
                document_errors = [
                    {"document": document["name"], "path": document["path"], **error}
                    for document in documents
                    for error in document.get("errors", [])
                ]
                self.send_json(
                    HTTPStatus.OK,
                    {
                        "dpi": dpi,
                        "documents": documents,
                        "totalPages": sum(document["pages"] for document in documents),
                        "renderedPages": sum(document["renderedPages"] for document in documents),
                        "errors": document_errors,
                        "renderedAt": datetime.now().isoformat(timespec="seconds"),
                    },
                )
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/dwg/model-page":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                page = int(body.get("page") or 1)
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not raw_file:
                    raise ValueError("Не выбран DWG-файл для отображения")
                pdf_path, convert_cache_hit = dwg_to_model_pdf(Path(raw_file))
                payload = render_pdf_page(pdf_path, page=page, dpi=dpi, page_timeout_seconds=DWG_MODEL_PAGE_TIMEOUT_SECONDS)
                payload["sourcePath"] = raw_file
                payload["sourceType"] = "DWG"
                payload["convertedPdfPath"] = str(pdf_path)
                payload["convertCacheHit"] = convert_cache_hit
                self.send_json(HTTPStatus.OK, payload)
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/excel/workbook":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                if not raw_file:
                    raise ValueError("Не выбран Excel-файл для отображения")
                self.send_json(HTTPStatus.OK, excel_workbook_preview(Path(raw_file)))
            except (ValueError, FileNotFoundError, RuntimeError, OSError, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/excel/sheet":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                sheet_index = int(body.get("sheetIndex", 0))
                sheet = excel_sheet_preview(Path(raw_file), sheet_index)
                self.send_json(HTTPStatus.OK, sheet)
            except (ValueError, IndexError, FileNotFoundError, RuntimeError, OSError, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/excel/render":
            try:
                body = self.read_json()
                raw_files = body.get("files", [])
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not isinstance(raw_files, list) or not raw_files:
                    raise ValueError("Не выбраны Excel-файлы для отображения")
                if len(raw_files) > 10:
                    raise ValueError("За один раз пока можно отрендерить не больше 10 Excel-файлов")
                documents = [render_excel(Path(str(file_path)), dpi=dpi) for file_path in raw_files]
                document_errors = [
                    {"document": document["name"], "path": document["path"], **error}
                    for document in documents
                    for error in document.get("errors", [])
                ]
                self.send_json(
                    HTTPStatus.OK,
                    {
                        "dpi": dpi,
                        "documents": documents,
                        "totalPages": sum(document["pages"] for document in documents),
                        "renderedPages": sum(document["renderedPages"] for document in documents),
                        "errors": document_errors,
                        "renderedAt": datetime.now().isoformat(timespec="seconds"),
                    },
                )
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/excel/page":
            try:
                body = self.read_json()
                raw_file = str(body.get("file", "")).strip()
                page = int(body.get("page") or 1)
                dpi = int(body.get("dpi") or DEFAULT_PDF_DPI)
                if dpi < 72 or dpi > 600:
                    raise ValueError("DPI должен быть в диапазоне 72-600")
                if not raw_file:
                    raise ValueError("Не выбран Excel-файл для отображения")
                pdf_path, convert_cache_hit = excel_to_pdf(Path(raw_file))
                payload = render_pdf_page(pdf_path, page=page, dpi=dpi)
                payload["sourcePath"] = raw_file
                payload["sourceType"] = file_extension(Path(raw_file))
                payload["convertedPdfPath"] = str(pdf_path)
                payload["convertCacheHit"] = convert_cache_hit
                self.send_json(HTTPStatus.OK, payload)
            except (ValueError, FileNotFoundError, RuntimeError, OSError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        if parsed.path == "/api/open-file":
            started = time.perf_counter()
            raw_file = ""
            target: Path | None = None
            try:
                body = self.read_json()
                raw_file = str(body.get("path", "")).strip()
                if not raw_file:
                    raise ValueError("Не выбран файл для открытия")
                target = Path(raw_file).expanduser()
                if not target.exists() or not target.is_file():
                    raise FileNotFoundError(f"Файл не найден: {target}")
                opened_path = target
                if os.name == "nt":
                    if target.suffix.casefold() == ".dwg":
                        opened_path = open_dwg_for_review(target)
                    else:
                        os.startfile(str(target))  # type: ignore[attr-defined]
                else:
                    subprocess.Popen(["xdg-open", str(target)])
                mode = "read-only-source" if target.suffix.casefold() == ".dwg" else "native"
                append_native_open_log(
                    {
                        "at": datetime.now().isoformat(timespec="seconds"),
                        "status": "requested",
                        "extension": target.suffix.casefold(),
                        "sourcePath": str(target),
                        "openedPath": str(opened_path),
                        "mode": mode,
                        "elapsedMs": round((time.perf_counter() - started) * 1000),
                    }
                )
                self.send_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "path": str(target),
                        "openedPath": str(opened_path) if target.suffix.casefold() == ".dwg" else str(target),
                        "mode": mode,
                    },
                )
            except (ValueError, FileNotFoundError, OSError, RuntimeError, subprocess.TimeoutExpired, json.JSONDecodeError) as error:
                append_native_open_log(
                    {
                        "at": datetime.now().isoformat(timespec="seconds"),
                        "status": "error",
                        "extension": target.suffix.casefold() if target else Path(raw_file).suffix.casefold(),
                        "sourcePath": str(target) if target else raw_file,
                        "error": str(error),
                        "elapsedMs": round((time.perf_counter() - started) * 1000),
                    }
                )
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})
            return

        self.send_json(HTTPStatus.NOT_FOUND, {"error": "Маршрут не найден"})

    def serve_cache(self, request_path: str) -> None:
        relative = unquote(request_path).removeprefix("/cache/").lstrip("/")
        target = (RUNTIME_DIR / "cache" / relative).resolve()

        try:
            target.relative_to((RUNTIME_DIR / "cache").resolve())
        except ValueError:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return

        if not target.exists() or not target.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        body = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "private, max-age=31536000, immutable")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def serve_static(self, request_path: str) -> None:
        relative = "index.html" if request_path in ("", "/") else unquote(request_path).lstrip("/")
        target = (FRONTEND_DIR / relative).resolve()

        try:
            target.relative_to(FRONTEND_DIR.resolve())
        except ValueError:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return

        if not target.exists() or not target.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        content_type = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        if content_type.startswith("text/") or content_type in {"application/javascript", "text/javascript"}:
            content_type = f"{content_type}; charset=utf-8"
        body = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run F-Engineering Launcher v3 local server.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8780)
    args = parser.parse_args()

    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    (RUNTIME_DIR / "cache").mkdir(exist_ok=True)
    PDF_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    MANIFESTS_DIR.mkdir(exist_ok=True)
    (RUNTIME_DIR / "logs").mkdir(exist_ok=True)

    server = ThreadingHTTPServer((args.host, args.port), LauncherHandler)
    print(f"F-Engineering Launcher v3: http://{args.host}:{args.port}/", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
